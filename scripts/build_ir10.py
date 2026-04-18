"""Build an IR10 Financial Statements Summary workbook from a consolidated context JSON.

Usage:
    python scripts/build_ir10.py --context <ir10-context.json> --output <IR10-filled.xlsx> [--config <ir10_config.json>]

Prints a single `RESULT: {...}` line to stdout on success (or error) for the orchestrating agent to parse.
"""
import argparse
import json
import sys
import traceback
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.comments import Comment


REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_CONFIG = REPO_ROOT / "guides" / "ir10_config.json"


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Build an IR10 workbook from a context JSON.")
    p.add_argument("--context", required=True, help="Path to ir10-context.json.")
    p.add_argument("--output", required=True, help="Path for the output .xlsx.")
    p.add_argument("--config", default=str(DEFAULT_CONFIG), help="Path to ir10_config.json.")
    return p.parse_args()


def load_json(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def build(ctx: dict, config: dict, output_path: Path) -> dict:
    ir10_form_url = config["ir10_form_url"]
    ir10_guide_url = config["ir10_guide_url"]
    depr_url = config["depreciation_tool_url"]

    taxpayer = ctx["taxpayer"]
    if not taxpayer.get("name") or not taxpayer.get("ird_number"):
        raise ValueError("taxpayer.name and taxpayer.ird_number are required")

    expenses = ctx.get("expenses", [])
    capital_items = ctx.get("capital_items", [])
    koinly_reports = ctx.get("koinly_reports", [])
    user_overrides = ctx.get("user_overrides", [])

    wb = Workbook()

    # ---------------- Sheet 2: Expenses ----------------
    exp_ws = wb.create_sheet("Expenses")
    exp_ws["A1"] = "EXPENSES"
    exp_ws["A1"].font = Font(bold=True, size=14)
    headers = ["Date", "Inv", "Seller", "Description", "Amount", "Type"]
    for i, h in enumerate(headers, start=1):
        c = exp_ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="B0C4DE")

    row = 3
    for e in expenses:
        r = e["result"]
        inv = r["invoice"]
        exp_ws.cell(row=row, column=1, value=inv.get("date"))
        exp_ws.cell(row=row, column=2, value=inv.get("number"))
        exp_ws.cell(row=row, column=3, value=inv.get("vendor"))
        exp_ws.cell(row=row, column=4, value=inv.get("description"))
        exp_ws.cell(row=row, column=5, value=r.get("nzd_amount"))
        exp_ws.cell(row=row, column=6, value=r.get("ir10_category"))
        row += 1
    exp_last = row - 1  # last data row (== 2 when empty)

    for col, width in zip("ABCDEF", [12, 16, 28, 45, 12, 36]):
        exp_ws.column_dimensions[col].width = width

    # ---------------- Sheet 3: Depreciation ----------------
    dep_ws = wb.create_sheet("Depreciation")
    dep_ws["A1"] = "Straight Line Depreciation"
    dep_ws["A1"].font = Font(bold=True, size=14)
    bal_year = ctx.get("balance_date", "2025-03-31")[:4]
    dep_headers = ["Date", "Inv", "Seller", "Description", "Opening Amount",
                   "Depreciation %", "Months Owned EOY", f"Depreciation Amount EOY Mar {bal_year}",
                   "Closing Amount", "IR10 Box"]
    for i, h in enumerate(dep_headers, start=1):
        c = dep_ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="B0C4DE")

    row = 3
    for cap in capital_items:
        r = cap["result"]
        inv = r["invoice"]
        dep_ws.cell(row=row, column=1, value=r.get("date_acquired"))
        dep_ws.cell(row=row, column=2, value=inv.get("number"))
        dep_ws.cell(row=row, column=3, value=inv.get("vendor"))
        dep_ws.cell(row=row, column=4, value=inv.get("description"))
        dep_ws.cell(row=row, column=5, value=r.get("opening_amount"))
        dep_ws.cell(row=row, column=6, value=r.get("depreciation_rate"))
        dep_ws.cell(row=row, column=7, value=f"=CHOOSE(MONTH(A{row}),3,2,1,12,11,10,9,8,7,6,5,4)")
        dep_ws.cell(row=row, column=8, value=f"=IF(E{row}<1000,E{row},E{row}*F{row}*(G{row}/12))")
        dep_ws.cell(row=row, column=9, value=f"=E{row}-H{row}")
        dep_ws.cell(row=row, column=10, value=r.get("ir10_box"))
        row += 1
    dep_last_data = row - 1  # last data row; == 2 when no capital items

    # TOTAL row
    total_row = row
    dep_ws.cell(row=total_row, column=7, value="TOTAL").font = Font(bold=True)
    if capital_items:
        dep_ws.cell(row=total_row, column=8, value=f"=SUM(H3:H{dep_last_data})").font = Font(bold=True)
        dep_ws.cell(row=total_row, column=9, value=f"=SUM(I3:I{dep_last_data})").font = Font(bold=True)
    else:
        # No data rows — write literal zeros so IR10 formulas resolve cleanly.
        dep_ws.cell(row=total_row, column=8, value=0).font = Font(bold=True)
        dep_ws.cell(row=total_row, column=9, value=0).font = Font(bold=True)

    for col, width in zip("ABCDEFGHIJ", [12, 14, 22, 30, 14, 14, 16, 22, 14, 10]):
        dep_ws.column_dimensions[col].width = width

    # ---------------- Sheet 1: IR10 ----------------
    # `wb.active` is the default sheet at index 0; renaming keeps [IR10, Expenses, Depreciation] order.
    ir10 = wb.active
    ir10.title = "IR10"

    ir10["A1"] = "IR10"
    ir10["A1"].font = Font(bold=True, size=16)
    ir10["B1"] = "Financial statements summary"
    ir10["B1"].font = Font(bold=True, size=14)
    ir10["F1"] = ir10_form_url
    ir10["F2"] = ir10_guide_url

    ir10["A4"] = "Your full name"
    ir10["B4"] = taxpayer["name"]
    ir10["A5"] = "Your IRD Number"
    ir10["B5"] = taxpayer["ird_number"]

    ir10["C6"] = "Box"
    ir10["E6"] = "Notes"
    ir10["F6"] = "Reference"
    for cell in ("C6", "E6", "F6"):
        ir10[cell].font = Font(bold=True)

    # Row layout: (row, A, B, box, D)
    rows_spec = [
        (8,  "Multiple activity indicator", None, 1, "No"),
        (10, "Profit and loss statement", None, None, None),
        (12, "Gross income from", "Sales and/or services", 2, 0),
        (14, "Cost of goods sold", "Opening stock (include work in progress)", 3, 0),
        (15, None, "Purchases", 4, 0),
        (16, None, "Closing stock (include work in progress)", 5, 0),
        (18, "Gross profit", "(if a loss, put a minus sign in the last box)", 6, "=D12-D14+D15-D16"),
        (20, "Other gross income", "Interest received", 7, 0),
        (21, None, "Dividends received", 8, 0),
        (22, None, "Rental, lease and licence income", 9, 0),
        (23, None, "Other income", 10, 0),
        (25, "Total income", "Add up all income entered in Boxes 6 to 10", 11, "=SUM(D18:D23)"),
        (27, "Expenses (as per financial statements)", "Bad debts", 12, '=SUMIF(Expenses!F:F,B27,Expenses!E:E)'),
        (28, None, "Accounting depreciation and amortisation", 13, f'=SUMIF(Expenses!F:F,B28,Expenses!E:E)+SUM(Depreciation!H3:H{dep_last_data})'),
        (29, None, "Insurance (exclude ACC levies)", 14, '=SUMIF(Expenses!F:F,B29,Expenses!E:E)'),
        (30, None, "Interest expense", 15, '=SUMIF(Expenses!F:F,B30,Expenses!E:E)'),
        (31, None, "Professional and consulting fees", 16, '=SUMIF(Expenses!F:F,B31,Expenses!E:E)'),
        (32, None, "Rates", 17, '=SUMIF(Expenses!F:F,B32,Expenses!E:E)'),
        (33, None, "Rental, lease and licence payments", 18, '=SUMIF(Expenses!F:F,B33,Expenses!E:E)'),
        (34, None, "Repairs and maintenance", 19, '=SUMIF(Expenses!F:F,B34,Expenses!E:E)'),
        (35, None, "Research and development", 20, '=SUMIF(Expenses!F:F,B35,Expenses!E:E)'),
        (36, None, "Associated persons' remuneration", 21, '=SUMIF(Expenses!F:F,B36,Expenses!E:E)'),
        (37, None, "Salaries and wages paid to employees", 22, '=SUMIF(Expenses!F:F,B37,Expenses!E:E)'),
        (38, None, "Contractor and sub-contractor payments", 23, '=SUMIF(Expenses!F:F,B38,Expenses!E:E)'),
        (39, None, "Other expenses", 24, '=SUMIF(Expenses!F:F,B39,Expenses!E:E)'),
        (41, "Total expenses", "Add up all expenses entered in Boxes 12 to 24", 25, "=SUM(D27:D39)"),
        (43, "Exceptional items", "(if there is a negative amount, put a minus sign in the last box)", 26, 0),
        (45, "Net profit/loss before tax", "Box 11 less Box 25, add Box 26", 27, "=D25-D41+D43"),
        (47, "Tax adjustments", "(if there is a negative amount, put a minus sign in the last box)", 28, 0),
        (49, "Current year taxable profit/loss", None, 29, "=D45+D47"),
        (51, "Balance sheet items", None, None, None),
        (53, "Current assets (as at balance date)", "Accounts receivable (debtors)", 30, 0),
        (54, None, "Cash and deposits", 31, 0),
        (55, None, "Other current assets", 32, 0),
        (57, "Fixed assets (closing accounting value)", "Vehicles", 33, "=SUMIF(Depreciation!J:J,33,Depreciation!I:I)"),
        (58, None, "Plant and machinery", 34, "=SUMIF(Depreciation!J:J,34,Depreciation!I:I)"),
        (59, None, "Furniture and fittings", 35, "=SUMIF(Depreciation!J:J,35,Depreciation!I:I)"),
        (60, None, "Land", 36, "=SUMIF(Depreciation!J:J,36,Depreciation!I:I)"),
        (61, None, "Buildings", 37, "=SUMIF(Depreciation!J:J,37,Depreciation!I:I)"),
        (62, None, "Other fixed assets", 38, "=SUMIF(Depreciation!J:J,38,Depreciation!I:I)"),
        (64, "Other non-current assets (as at balance date)", "Intangibles", 39, 0),
        (65, None, "Shares/ownership interests", 40, 0),
        (66, None, "Term deposits", 41, 0),
        (67, None, "Other non-current assets", 42, 0),
        (69, "Total assets", "Add up all assets entered in Boxes 30 to 42", 43, "=SUM(D53:D67)"),
        (71, "Current liabilities (as at balance date)", "Provisions", 44, 0),
        (72, None, "Accounts payable (creditors)", 45, 0),
        (73, None, "Current loans", 46, 0),
        (74, None, "Other current liabilities", 47, 0),
        (75, "Total current liabilities", "Add up all liabilities entered in Boxes 44 to 47", 48, "=SUM(D71:D74)"),
        (77, "Non-current liabilities (as at balance date)", None, 49, 0),
        (79, "Total liabilities", "Add Box 48 to Box 49", 50, "=D75+D77"),
        (81, "Owners' equity", "(if in debit, put a minus sign in the last box)", 51, "=D69-D79"),
        (83, "Other information", "Tax depreciation", 52, "=D28"),
        (84, None, "Untaxed realised gains/receipts", 53, 0),
        (85, None, "Additions to fixed assets", 54, f'=SUMIF(Expenses!F:F,"Accounting depreciation and amortisation",Expenses!E:E)+SUM(Depreciation!E3:E{dep_last_data})'),
        (86, None, "Disposals of fixed assets", 55, 0),
        (87, None, "Dividends paid", 56, 0),
        (88, None, "Drawings", 57, 0),
        (89, None, "Current account year-end balances", 58, 0),
        (90, None, "Tax-deductible loss on disposal of fixed assets", 59, 0),
    ]

    # Koinly aggregation (sum across all reports)
    koinly_sources = [k.get("source_file", "") for k in koinly_reports]
    income_total = sum((k.get("result", {}).get("income", {}) or {}).get("income_summary_total", 0) or 0 for k in koinly_reports)
    cap_gains = sum((k.get("result", {}).get("income", {}) or {}).get("capital_gains_net", 0) or 0 for k in koinly_reports)
    other_gains = sum((k.get("result", {}).get("income", {}) or {}).get("other_gains_net", 0) or 0 for k in koinly_reports)
    eoy_total = sum((k.get("result", {}).get("end_of_year_balances", {}) or {}).get("total_value", 0) or 0 for k in koinly_reports)

    # Box → (cell value, sources, notes) overrides driven by Koinly
    koinly_overrides = {}
    if koinly_reports:
        koinly_overrides[2] = (income_total, koinly_sources, "Koinly/Income Summary/Total")
        koinly_overrides[28] = (f"={cap_gains}+{other_gains}", koinly_sources, "Koinly/Capital gains net + Other gains net")
        koinly_overrides[39] = (eoy_total, koinly_sources, "Koinly/End of year balances")

    # Which boxes default to "User to provide" when left at literal 0
    user_to_provide_boxes = {26, 30, 31, 32, 39, 40, 41, 42, 44, 45, 46, 47, 49, 53, 55, 56, 57, 58, 59}

    # Build box → row map from rows_spec
    box_to_row = {box: r for (r, _a, _b, box, _d) in rows_spec if box is not None}

    # User overrides (explicit {"box": N, "value": V}) take precedence over everything
    override_notes = {}  # box → note string
    override_sources = {}  # box → list of source strings
    for ov in user_overrides:
        box = ov.get("box")
        if box is None or box not in box_to_row:
            continue
        override_notes[box] = ov.get("note") or ov.get("source") or "User override"
        override_sources[box] = [ov.get("source")] if ov.get("source") else []

    # Write rows
    boxes_filled = []
    boxes_user_to_provide = []
    for (r, a, b, box, d) in rows_spec:
        if a is not None:
            ir10.cell(row=r, column=1, value=a).font = Font(bold=True)
        if b is not None:
            ir10.cell(row=r, column=2, value=b)
        if box is not None:
            ir10.cell(row=r, column=3, value=box)

        # Resolve cell value in precedence order: user override > koinly override > rows_spec default.
        value = d
        note_for_e = None
        sources_for_box = []
        if box is not None and box in koinly_overrides:
            value, sources_for_box, note_for_e = koinly_overrides[box]
        user_ov_applied = False
        if box is not None:
            for ov in user_overrides:
                if ov.get("box") == box:
                    value = ov.get("value")
                    user_ov_applied = True
                    if override_notes.get(box):
                        note_for_e = override_notes[box]
                    if override_sources.get(box):
                        sources_for_box = override_sources[box]
                    break

        if value is not None:
            cell = ir10.cell(row=r, column=4, value=value)
            if isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right")

        # Write notes column for koinly/override-driven rows
        if note_for_e:
            ir10.cell(row=r, column=5, value=note_for_e)

        # Track outputs
        if box is not None:
            is_formula = isinstance(value, str) and value.startswith("=")
            is_literal_zero = isinstance(value, (int, float)) and value == 0
            if is_formula:
                boxes_filled.append({
                    "box": box,
                    "value": value,
                    "sources": sources_for_box,
                    "notes": note_for_e or "",
                })
            elif value is not None and not is_literal_zero:
                boxes_filled.append({
                    "box": box,
                    "value": value,
                    "sources": sources_for_box,
                    "notes": note_for_e or ("User override" if user_ov_applied else ""),
                })
            elif is_literal_zero and box in user_to_provide_boxes and not user_ov_applied:
                boxes_user_to_provide.append(box)
                ir10.cell(row=r, column=4).comment = Comment("User to provide", "ir10-agent")

    # Static reference annotations
    ir10["E28"] = "Fixed assets under $1000 + depreciation"
    ir10["F28"] = depr_url

    # Column widths
    for col, width in zip("ABCDEF", [40, 42, 6, 18, 42, 40]):
        ir10.column_dimensions[col].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "status": "ok",
        "excel_path": str(output_path),
        "sheets": ["IR10", "Expenses", "Depreciation"],
        "expenses_row_count": len(expenses),
        "capital_items_row_count": len(capital_items),
        "boxes_filled": boxes_filled,
        "boxes_user_to_provide": sorted(boxes_user_to_provide),
    }


def main() -> int:
    args = parse_args()
    try:
        ctx = load_json(Path(args.context))
        config = load_json(Path(args.config))
        result = build(ctx, config, Path(args.output))
    except Exception as exc:  # noqa: BLE001 — surface all errors via RESULT line
        err = {"status": "error", "message": f"{type(exc).__name__}: {exc}"}
        sys.stderr.write(traceback.format_exc())
        print("RESULT: " + json.dumps(err))
        return 1
    print("RESULT: " + json.dumps(result))
    return 0


if __name__ == "__main__":
    sys.exit(main())
